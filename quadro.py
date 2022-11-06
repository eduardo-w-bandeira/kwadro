'''
# *quadro* is a rudimentary ORM for xlsx files

# First of all, install openpyxl, once this layer is written on the top of it

# Example of usage:

from quadro import BaseTable, Column, Board

# Define a table, i.e. a sheet-class
class Clients(BaseTable):
    __title__ = 'Clients'
    name = Column(1)  
    phone = Column(2)
    address = Column('C') # or Column(3)
    country = Column(4)


# Define your file
board = Board('my-file.xlsx') # new file or existing one

# Create the sheet, if not created
board.create_sheet(Clients)

# Or force a new one
board.create_sheet(Clients, force=True)

# Add an entry at first empty row
entry = Clients(
    name='Joshua King', phone=123,
    address='80 Bla St, Birmingham',
    country='Australia')

board.add(entry)

# If one wants to get the row
print(entry._row) # Outputs: 1

# Find an entry
entry = board.find(Clients, name='Joshua King',
    address='80 Bla St, Birmingham')

print(entry.phone) # Outputs: 123

# find_all() yields matches
entries = board.find_all(Clients, country='Australia')

# If one wants to access openpyxl-Worksheet object
worksheet = entry._worksheet

# Saving
board.save('new-or-same-file.xlsx')
'''

import os

from openpyxl import Workbook, load_workbook

class Column:

    def __init__(self, letter_or_number):
        ''':param letter_or_number: column letter or
        column number'''
        letter = letter_or_number
        if isinstance(letter_or_number, int):
            n = letter_or_number
            letter = ''
            while n > 0:
                n, remainder = divmod(n-1, 26)
                letter = chr(65 + remainder) + letter
        self.colletter = letter


class BaseTable:
    _worksheet = None

    def __init__(self, **kwargs):
        ''':param kwargs: These are column values,
        wich will be assigned automatically.'''
        colname_colletter_map = {}
        self._row = None
        for key, value in self.__class__.__dict__.items():
            if isinstance(value, Column):
                colname_colletter_map[key] = value.colletter
                self.__setattr__(key, value)
        # Sort colname_colletter_map by values
        self._colname_colletter_map = dict(sorted(
            colname_colletter_map.items(), key=lambda item: item[1]))
        for key, value in kwargs.items():
            self.__setattr__(key, value)

    def __setattr__(self, name, value):
        object.__setattr__(self, name, value)
        self._assign_cell_value(name, value)

    def __getattribute__(self, name):
        ws = object.__getattribute__(self, '_worksheet')
        if not ws:
            return object.__getattribute__(self, name)
        colname_colletter = object.__getattribute__(
            self, '_colname_colletter_map')
        if name not in colname_colletter:
            return object.__getattribute__(self, name)
        value = self._get_cell_value(name)
        object.__setattr__(self, name, value) # update __dict__
        return value

    def _associate_worksheet_and_row(self, worksheet, row):
        self._worksheet = worksheet
        self._row = row

    def _get_cell(self, colname):
        colletter = self._colname_colletter_map[colname]
        return self._worksheet[f'{colletter}{self._row}']

    def _assign_cell_value(self, colname, value):
        if self._worksheet and colname in self._colname_colletter_map:
            if not isinstance(value, Column): # Assures the value was provided
                self._get_cell(colname).value = value

    def _get_cell_value(self, colname):
        return self._get_cell(colname).value


class Board:
    'One can access openpyxl.Workbook through Board.workbook'

    def __init__(self, file=None, force_new=False):
        '''
        :param file: New or existing file.
        If new, default woksheet will be removed.
        :type file: str or pathlib.Path object.
        :param force_new: If True, pre-existing file will be replaced.
        '''
        self.file = file
        self._wsrow_entry_map = {}
        if not file or force_new or not os.path.exists(file):
            self.workbook = Workbook()
            # delete default worksheet
            self.workbook.remove(self.workbook.active)
        else:
            self.workbook = load_workbook(file)

    def create_sheet(self, table, index=None, force_new=False):
        '''
        Creates the sheet if not already created.
        :type table: BaseTable (class).
        :param force_new: if True, pre-existing sheet will be deleted
        before creating new.
        '''
        title = table.__title__
        if title in self.workbook.sheetnames and force_new:
            self.workbook.remove(self.workbook[title])
        if title not in self.workbook.sheetnames:
            self.workbook.create_sheet(title, index)

    def create_and_add(self, entry, index=None, force_new=False):
        '''
        Creates the sheet and add.
        :type entry: a table instance
        '''
        self.create_sheet(entry.__class__, index, force_new)
        self.add(entry)

    def get(self, table, row):
        '''Provides the entry by row'''
        ws = self.workbook[table.__title__]
        if (ws, row) in self._wsrow_entry_map:
            return self._wsrow_entry_map[(ws, row)]
        entry = table()
        entry._associate_worksheet_and_row(ws, row)
        self._wsrow_entry_map[(ws, row)] = entry
        return entry

    def find(self, table_or_entries, **kwargs):
        '''
        :type table_or_entries: a table (class) or a
        list of entries.
        :returns: first match entry or None
        '''
        if isinstance(table_or_entries, list):
            table = table_or_entries[0].__class__
            searching_rows = [entry._row for entry in table_or_entries]
        else:
            table = table_or_entries
            empty_row = self.last_row(table) + 1
            searching_rows = range(1, empty_row)
        for row in self._find_rows(table, searching_rows, **kwargs):
            return self.get(table, row)
        return None # just explicting

    def find_all(self, table_or_entries, **kwargs):
        '''
        :type table_or_entries: a table (class) or a list of entries.
        :param kwargs: if not provided, all entries will be returned.
        :returns: a list of match entries
        '''
        if isinstance(table_or_entries, list):
            if not kwargs:
                return table_or_entries
            table = table_or_entries[0].__class__
            searching_rows = [entry._row for entry in table_or_entries]
            match_rows = self._find_rows(table, searching_rows, **kwargs)
        else:
            table = table_or_entries
            empty_row = self.last_row(table) + 1
            searching_rows = range(1, empty_row)
            if not kwargs:
                match_rows = searching_rows
            else:
                match_rows = self._find_rows(table, searching_rows, **kwargs)
        for row in match_rows:
            yield self.get(table, row)

    def _find_rows(self, table, searching_rows, **kwargs):
        for row in searching_rows:
            entry = self.get(table, row)
            matches = []
            for colname, value in kwargs.items():
                if entry._get_cell_value(colname) != value:
                    break
                matches.append(True)
                if len(matches) == len(kwargs):
                    yield row
                    break

    def last_row(self, table):
        ws = self.workbook[table.__title__]
        max_row = ws.max_row
        if max_row == 1: # Because ws.max_row is never < 1
            first = self.get(table, 1)
            values = [
                first._get_cell_value(colname) for colname in
                first._colname_colletter_map.keys()]
            if not any(values):
                max_row = 0
        return max_row

    def add(self, entry):
        ''':type entry: a table instance'''
        ws = self.workbook[entry.__title__]
        row = self.last_row(entry.__class__) + 1
        entry._associate_worksheet_and_row(ws, row)
        for key in entry._colname_colletter_map.keys():
            value = entry.__dict__[key]
            entry._assign_cell_value(key, value)
        self._wsrow_entry_map[(ws, row)] = entry

    def remove(self, entry):
        entry._worksheet.delete_entries(entry._row)
        ws = self.workbook[entry.__title__]
        del self._wsrow_entry_map[(ws, entry._row)]

    def has_table(self, table):
        if table.__title__ in self.workbook.sheetnames:
            return True
        return False

    def save(self, file):
        self.workbook.save(file)