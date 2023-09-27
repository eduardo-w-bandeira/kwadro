"""
*kwadro* is basic ORM for Excel files (xlsx),
that works with openpyxl package.

You will find examples of use in "README.md".
"""

__all__ = ['Table', 'Column', 'Board']

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter


class Column:

    def __init__(self, letter_or_number):
        """
        Args:
            letter_or_number (str or int): Column letter or column number.
                The column number starts at 1.
        """
        letter = letter_or_number
        if isinstance(letter_or_number, int):
            letter = get_column_letter(letter_or_number)
        self.colletter = letter.upper()


class Table:
    """
    ```
    class Employees(Table):
        __title__ = "Employees"
        name = Column("A") # Or Column(1) 
        birth = Column("B")
        phone = Column("c") # It's case insensitive
        address = Column(4) # You can use numbers instead
        country = Column("E")
    ```
    That is, define a Table-class for every sheet you want to work with.
    Your class should inherit from this class.

    Special variable:
        __title__ (str): When defining your class, you must use
            this special variable with the name of the worksheet.
    """
    _worksheet = None

    def __init__(self, **column_values):
        """
        Args:
            column_values: `employee = Employees(name="John Doe", phone=7654321)`.
                It's optional, you may add the values after instancing it:
                ```
                employee = Employees()
                employee.name = "John Doe"
                employee.phone = 7654321
                ```
        """
        colname_colletter_map = {}
        self._row = None
        self._board = None
        for key, value in self.__class__.__dict__.items():
            if isinstance(value, Column):
                colname_colletter_map[key] = value.colletter
                self.__setattr__(key, value)
        # Sort colname_colletter_map by values
        self._colname_colletter_map = dict(sorted(
            colname_colletter_map.items(), key=lambda item: item[1]))
        for key, value in column_values.items():
            self.__setattr__(key, value)

    def __setattr__(self, name, value):
        object.__setattr__(self, name, value)
        self._assign_cell_value(name, value)

    def __getattribute__(self, name):
        ws = object.__getattribute__(self, "_worksheet")
        if not ws:
            return object.__getattribute__(self, name)
        colname_colletter = object.__getattribute__(
            self, "_colname_colletter_map")
        if name not in colname_colletter:
            return object.__getattribute__(self, name)
        value = self._get_cell_value(name)
        object.__setattr__(self, name, value)  # update __dict__
        return value

    def _assign_internal_data(self, worksheet, row, board):
        self._worksheet = worksheet
        self._row = row
        self._board = board

    def _get_cell(self, colname):
        colletter = self._colname_colletter_map[colname]
        return self._worksheet[f"{colletter}{self._row}"]

    def _assign_cell_value(self, colname, value):
        if self._worksheet and colname in self._colname_colletter_map:
            if not isinstance(value, Column):  # Assures the value was provided
                self._get_cell(colname).value = value

    def _get_cell_value(self, colname):
        return self._get_cell(colname).value


class Board:
    """
    Definitions:
        table: A class derived from Table (sheet class).
        record: An instance of a Table (i.e. line entries).
    """

    def __init__(self, file=None):
        """
        Args:
            file (str or Pathlib.Path instance): The path of the file.
                If not provided, a new one will be created,
                and default woksheet will be removed.
        """
        self._wsrow_record_map = {}
        if not file:
            self._workbook = Workbook()
            # delete default worksheet
            self._workbook.remove(self._workbook.active)
        else:
            self._workbook = load_workbook(file)

    def create_sheet(self, table, index=None, force_new=False):
        """Creates the sheet if not already created (at an optional index).

        Args:
            table (Table): A Table-derived class.
            index (int): Optional position at which the sheet will
                be inserted. It starts at 0.
            force_new (bool): if True, pre-existing sheet
                will be deleted before creating new.

        Returns:
            None.
        """
        title = table.__title__
        if title in self._workbook.sheetnames and force_new:
            self._workbook.remove(self._workbook[title])
        if title not in self._workbook.sheetnames:
            self._workbook.create_sheet(title, index)

    def create_and_add(self, record, index=None, force_new=False):
        """Creates the sheet and adds the record to the file.

        Args:
            record: A Table-derived instance.
            index (int): Optional position at which the sheet will
                be inserted. It starts at 0.
            force_new (bool): if True, pre-existing sheet
                will be deleted before creating new.

        Returns:
            None.
        """
        self.create_sheet(record.__class__, index, force_new)
        self.add(record)

    def get_record(self, table, row):
        """Returns the record by its row number.

        Args:
            table (Table): A Table-derived class.
            row (int): The row number.

        Returns:
            The record object (i.e. the table instance).
        """
        ws = self._workbook[table.__title__]
        if (ws, row) in self._wsrow_record_map:
            return self._wsrow_record_map[(ws, row)]
        record = table()
        record._assign_internal_data(ws, row, self)
        self._wsrow_record_map[(ws, row)] = record
        return record

    def find(self, table_or_records, **kwargs):
        """
        Args:
            table_or_records (Table or list): Table-derived class or
                a list of records.

        Returns:
            The first matching record if found. Otherwise, None.
        """
        if isinstance(table_or_records, list):
            table = table_or_records[0].__class__
            searching_rows = [record._row for record in table_or_records]
        else:
            table = table_or_records
            empty_row = self.last_row(table) + 1
            searching_rows = range(1, empty_row)
        for row in self._find_rows(table, searching_rows, **kwargs):
            return self.get_record(table, row)
        return None  # just explicting

    def find_all(self, table_or_records, **column_values):
        """
        Args:
            table_or_records (Table or list): A Table-derived class or
                a list of records.
            column_values: `board.find(Employees, name="Bla-Bla-Bla", Country="Australia")`.
                These are filters. If not provided, all records will be retrieved.

        Yields:
            Matched records.
        """
        if isinstance(table_or_records, list):
            if not column_values:
                return table_or_records
            table = table_or_records[0].__class__
            searching_rows = [record._row for record in table_or_records]
            match_rows = self._find_rows(
                table, searching_rows, **column_values)
        else:
            table = table_or_records
            empty_row = self.last_row(table) + 1
            searching_rows = range(1, empty_row)
            if not column_values:
                match_rows = searching_rows
            else:
                match_rows = self._find_rows(
                    table, searching_rows, **column_values)
        for row in match_rows:
            yield self.get_record(table, row)

    def _find_rows(self, table, searching_rows, **column_values):
        for row in searching_rows:
            record = self.get_record(table, row)
            matches = []
            for colname, value in column_values.items():
                if record._get_cell_value(colname) != value:
                    break
                matches.append(True)
                if len(matches) == len(column_values):
                    yield row
                    break

    def last_row(self, table):
        """Gets the last row number that is not empty.

        Args:
            table (Table): A Table-derived class.

        Returns (int):
            The last row number that is not empty.
            If 0, it means the whole worksheet is empty.
        """
        ws = self._workbook[table.__title__]
        max_row = ws.max_row
        if max_row == 1:  # Because ws.max_row is never < 1
            first = self.get_record(table, 1)
            values = [
                first._get_cell_value(colname) for colname in
                first._colname_colletter_map.keys()]
            if not any(values):
                return 0  # The worksheet is empty.
        return max_row

    def add(self, record):
        """Adds the record to the first empty row.

        Args:
            record: A table instance.

        Returns:
            None.
        """
        ws = self._workbook[record.__title__]
        row = self.last_row(record.__class__) + 1
        record._assign_internal_data(ws, row, self)
        for key in record._colname_colletter_map.keys():
            value = record.__dict__[key]
            record._assign_cell_value(key, value)
        self._wsrow_record_map[(ws, row)] = record

    def remove(self, record):
        """Removes the record from the worksheet.

        Args:
            record: A table instance.

        Returns:
            None.
        """
        record._worksheet.delete_entries(record._row)
        ws = self._workbook[record.__title__]
        del self._wsrow_record_map[(ws, record._row)]

    def has_table(self, table):
        """Checks whether the Table exists.

        Args:
            table (Table): A Table-derived class.

        Returns (bool):
            True, if it has it. Otherwise, False.
        """
        if table.__title__ in self._workbook.sheetnames:
            return True
        return False

    def save(self, file):
        """Saves the workbook in the provided file.

        Args:
            file (str or Pathlib.Path instance): The path of XLSX file.
                It may be the same loaded file or a new one.

        Returns:
            None.
        """
        self._workbook.save(file)
